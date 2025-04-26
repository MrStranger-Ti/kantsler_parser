import logging
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from src import config
from src.handling.base.handler import FormatHandler
from src.handling.mixins import ProductsHandlerMixin
from src.utils.tqdm import (
    CREATING_EXCEL_CONFIG,
    FORMATTING_PRODUCTS_CONFIG,
    SAVING_EXCEL_CONFIG,
)
from src.handling.excel import formatters

log = logging.getLogger(__name__)


class ExcelColumn:
    def __init__(self, data: Any, num: int) -> None:
        self.data = data
        self.num = num


class ExcelHandler(FormatHandler, ProductsHandlerMixin):
    formatters_classes = [
        formatters.ConstantFieldsFormatter,
        formatters.DefaultFieldsFormatter,
        formatters.SKUFormatter,
        formatters.PriceFormatter,
        formatters.CategoryFormatter,
        formatters.ImagesFormatter,
        formatters.CharacteristicsFormatter,
        formatters.PackageFormatter,
        formatters.SizesFormatter,
    ]

    def __init__(
        self,
        *args,
        template_path: Path | str = config.TEMPLATE_PATH,
        sheet_name: str = config.SHEET_NAME,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.template_path: Path | str = template_path
        self.sheet_name: str = sheet_name
        self._workbook: Workbook | None = None
        self._sheet: Worksheet | None = None

    @contextmanager
    def process(self):
        self._workbook = self.get_workbook()
        self._sheet = self.get_sheet()

        yield

        with tqdm(**SAVING_EXCEL_CONFIG):
            self.save_workbook()
            self._workbook.close()

    def get_workbook(self) -> Workbook:
        return load_workbook(self.template_path)

    def get_sheet(self) -> Worksheet:
        try:
            sheet = self._workbook[self.sheet_name]
        except KeyError as exc:
            log.error(
                f"{str(exc)}. Sheet name {config.SHEET_NAME} not found. Set correct in config.ini."
            )
            raise exc

        return sheet

    def save_workbook(self) -> None:
        log.info("Saving file")
        try:
            self._workbook.save(config.RESULT_FILE_DIR)
        except PermissionError as exc:
            log.error(f"{str(exc)}. Close result excel file and try again.")
            raise exc

        log.info("Excel file has been saved")

    def get_cols(self, items: list[dict[str, str]]) -> list[list[ExcelColumn]]:
        return [
            [
                ExcelColumn(data=value, num=config.COLS[name])
                for name, value in item.items()
                if name in config.COLS
            ]
            for item in items
        ]

    def format(self, data: dict[str, list[dict[str, Any]]]) -> list[dict[str, str]]:
        log.info("Starting to format data.")
        products = data.get("data")

        formatted_products = []
        for product in tqdm(products or [], **FORMATTING_PRODUCTS_CONFIG):
            formatted_item = self.process_formatters(data=product)
            if formatted_item:
                formatted_products.append(formatted_item)

        sorted_formatted_products = self.sort_products_by_oldest(
            products=formatted_products
        )

        log.info("Data has been formatted")
        return sorted_formatted_products

    def handle(self, data: list[dict[str, str]]) -> None:
        log.info("Starting to prepare excel file")
        rows = self.get_cols(items=data)

        with self.process():
            for row_num, row in enumerate(
                tqdm(rows, **CREATING_EXCEL_CONFIG), start=config.START_ROW
            ):
                for col in row:
                    self._sheet.cell(row=row_num, column=col.num, value=col.data)

        self.save_products_ids(products=data)

        log.info("Excel file has been created")
